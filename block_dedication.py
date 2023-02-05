'''
Program to ingest student data and produce individual reports in
a spreadsheet
'''
from tkinter import filedialog, messagebox, Tk
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.utils import get_column_letter, datetime
from openpyxl.chart import BarChart, Reference
import pandas as pd
from datetime import timedelta
from statistics import mean
import json
import time
from functools import reduce


def select_workbook(window_title, initialpath):
    filetypes = (
        ("Excel Workbook", "*.xlsx"),
        ("Excel Macro-Enabled Workbook (code)", "*.xlsm"),
    )

    filename = filedialog.askopenfilename(
        title=window_title,
        initialdir=initialpath,
        filetypes=filetypes)

    if filename == "":
        messagebox.showerror("File Error", "Please, select a workbook")
        exit()
    else:
        return filename


def as_text(value):
    if value is None:
        return ""
    return str(value)


def column_fit_properly(worksheet):
    '''Tries to properly fit all olumns in a worksheet, however, it doesn't
    always work due to diferent font size. For better results use monospaced fonts'''
    for column_cells in worksheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(
            column_cells[0].column)].width = length


def seconds_to_str(time_value):
    '''Converts seconds to string in a single line'''
    return "%d:%02d:%02d.%03d" % \
        reduce(lambda ll, b: divmod(ll[0], b) + ll[1:],
               [(time_value*1000,), 1000, 60, 60])


def get_folder_std():
    '''Get values from JSON to make it faster to select the desired folder'''
    if os.path.exists('./base.json'):
        with open('base.json') as base_file:
            data = json.load(base_file)
            initial_location = data["path"]
    else:
        initial_location = ""
    return initial_location


def update_folder_std(updated_path):
    '''Update folder standard path in base.json file'''
    data = {}
    data['path'] = os.path.dirname(updated_path)
    with open('base.json', 'w') as base_file:
        json.dump(data, base_file)
    return None


def load_excel_properly(path):
    '''loads excel file properly depending on file type,
    includes settings for timeseries'''
    workbook_filetype = os.path.splitext(path)[1]
    if workbook_filetype == ".xlsm":
        workbook = load_workbook(
            path, read_only=False, keep_vba=True)
    else:
        workbook = load_workbook(
            path, read_only=False, keep_vba=False)
    workbook.iso_dates = True
    workbook.epoch = datetime.CALENDAR_MAC_1904
    return workbook


def add_styles(workbook):
    '''Add all styles used for this programm if they don't exist in the excel workbook'''
    header_style = NamedStyle(name="Header Style")
    header_style.font = Font(bold=True,
                             color='00FFFFFF')
    header_style.fill = PatternFill(fill_type="solid",
                                    start_color='FF000000',
                                    )
    if header_style.name not in workbook.style_names:
        workbook.add_named_style(header_style)


def clean_sheets(workbook, exceptions=None):
    '''Remove all sheets from a workbook, but the ones in exceptions'''
    for sheet in workbook.sheetnames:
        if sheet not in exceptions:
            del workbook[sheet]


def fill_timedelta_dict(worksheet):
    dictonary = {}
    for atividade in range(2, worksheet.max_row+1):
        atividade_title = worksheet.cell(atividade, 1).value
        time_delta = worksheet.cell(atividade, 2).value

        if time_delta == 0 or time_delta == "" or time_delta is None:
            messagebox.showerror("Value Error", "Insert time delta")
            exit()
        else:
            dictonary[atividade_title] = pd.to_timedelta(
                time_delta, unit="minute")
    return dictonary


def set_default_dict(key_list):
    '''Initialize values as 0 in a dictionary'''
    dictonary = {}
    for key in range(len(key_list)):
        key_name = key_list[key]
        dictonary.setdefault(key_name, pd.Timedelta(seconds=0))
    return dictonary


def calculate_dt(dataframe, sort_list, group_list, calculate_over, delta):
    dataframe = dataframe.sort_values(by=sort_list)
    dataframe[delta] = dataframe.groupby(
        group_list)[calculate_over].apply(lambda x: x.diff())
    dataframe[delta].fillna(value=pd.Timedelta(seconds=0), inplace=True)
    return dataframe


def fill_delta(dataframe, dictionary):
    for delta in dictionary:
        dataframe = calculate_dt(dataframe, sort_list=dictionary[delta]["sort_order"],
                                 group_list=dictionary[delta]["group_order"], calculate_over="time", delta=delta)
    return dataframe


start_time = time.time()

# hide tkinter window
Tk().withdraw()

# Select control workbook
control_location = select_workbook(
    window_title="Select Control Workbook",
    initialpath=get_folder_std()
)

update_folder_std(control_location)
os.chdir(os.path.dirname(control_location))
control_book = load_excel_properly(control_location)

gabarito_sheet = control_book["Gabarito"]

if gabarito_sheet.max_row == 1:
    messagebox.showerror("Value Error", "No activities in template")
    exit()

time_delta_dict = fill_timedelta_dict(worksheet=gabarito_sheet)
dedication_general_name = "Dedicação Geral Independente da Atividade estar no Gabarito"
time_delta_general = time_delta_dict[dedication_general_name]
time_delta_dict = {keys: time_delta_dict[keys]
                   for keys in time_delta_dict.keys() - {dedication_general_name}}
activity_list = list(time_delta_dict)
activity_list.sort()

sheets_to_keep = ["Gabarito"]
clean_sheets(workbook=control_book, exceptions=sheets_to_keep)

# Ingerir logs
source_location = select_workbook(
    window_title="Select Activity Logs",
    initialpath=os.getcwd())

source_book = load_excel_properly(source_location)
source_sheet_name = source_book.sheetnames[0]

source_df = pd.read_excel(
    source_location, sheet_name=source_sheet_name, parse_dates=["Hora"])

print(source_df.head())
# filtrar colunas uteis
atividades_df = source_df[["Hora", "Nome completo", "Contexto do Evento"]]
atividades_df.columns = ["time", "full_name", "activity"]
atividades_df.columns = atividades_df.columns.str.strip()

# Criar a lista de alunos
student_list = list(set(atividades_df["full_name"].to_numpy()))
student_list.sort()

# Set up dictionaries
activity_sum_dict = set_default_dict(key_list=activity_list)
dedication_specific_dict = set_default_dict(key_list=student_list)
dedication_general_dict = set_default_dict(key_list=student_list)

# Prepare specific dt and overall_dt
dt_dict = {
    "activity_dt": {
        "sort_order": ["full_name", "activity", "time"],
        "group_order": ["full_name", "activity"]},
    "overall_dt": {
        "sort_order": ["full_name", "time", "activity"],
        "group_order": ["full_name"]}}

atividades_df = fill_delta(dataframe=atividades_df, dictionary=dt_dict)

HEADER_STYLE = "Header Style"

# Para cada aluno, preencher os dados de cada atividade
for student in range(len(student_list)):
    student_name = student_list[student]
    student_sheet = control_book.create_sheet(title=student_name[:31].title())

    # headers
    student_sheet.cell(1, 1).value = "Atividade"
    student_sheet.cell(1, 2).value = "Dedicação (h:mm)"

    # get a slice of an student
    student_df = atividades_df[atividades_df["full_name"] == student_name][
        ["time", "activity", "activity_dt", "overall_dt"]]

    # initialize standard values
    dedicacao_especifica = pd.Timedelta(seconds=0)
    dedicacao_especifica_total = pd.Timedelta(seconds=0)
    dedicacao_geral = pd.Timedelta(seconds=0)

    # Loop para dedicacao especifica
    for activity in range(len(activity_list)):
        activity_rows = activity+2  # variavel para facilitar o loop
        activity_name = activity_list[activity]
        time_delta = time_delta_dict[activity_name]
        # get the slice for an activity
        activity_df = student_df[
            (student_df["activity"] == activity_name)
            & (student_df["activity_dt"] <= time_delta)][
            ["time", "activity_dt"]]

        dedicacao_especifica = activity_df["activity_dt"].sum()
        activity_sum_dict[activity_name] += dedicacao_especifica
        dedicacao_especifica_total = dedicacao_especifica_total+dedicacao_especifica
        student_sheet.cell(activity_rows, 1).value = activity_list[activity]
        student_sheet.cell(activity_rows, 2).value = dedicacao_especifica
        student_sheet.cell(activity_rows, 2).number_format = "[h]:mm"

    dedicacao_geral = student_df[student_df["overall_dt"] <= time_delta_general][
        "overall_dt"].sum()

    # headers
    student_sheet.cell(activity_rows+1, 1).value = "Total das atividades acima"
    student_sheet.cell(activity_rows+2, 1).value = dedication_general_name
    # values
    student_sheet.cell(activity_rows+1, 2).value = dedicacao_especifica_total
    student_sheet.cell(activity_rows+2, 2).value = dedicacao_geral
    # fill dictionaries
    dedication_specific_dict[student_name] += dedicacao_especifica_total
    dedication_general_dict[student_name] += dedicacao_geral
    # formatting
    student_sheet.cell(1, 1).style = HEADER_STYLE
    student_sheet.cell(1, 2).style = HEADER_STYLE
    student_sheet.cell(activity_rows+1, 2).number_format = "[h]:mm"
    student_sheet.cell(activity_rows+2, 2).number_format = "[h]:mm"

    column_fit_properly(student_sheet)

# Preencher a tabela de resultados

summary_sheet = control_book.create_sheet(title="Resultado", index=1)

summary_header = ["Aluno",
                  "Tempo Dedicado Total Específico (h:mm)",
                  "Tempo Dedicado Total Relativo (%)",
                  "Tempo Dedicado Geral Absoluto (h:mm)",
                  "Tempo Dedicado Geral Relativo (%)"]

for header in range(len(summary_header)):
    column_index = header+1
    summary_sheet.cell(1, column_index).value = summary_header[header]
    summary_sheet.cell(1, column_index).style = HEADER_STYLE

dedication_specific_max = max(dedication_specific_dict.values())
dedication_general_max = max(dedication_general_dict.values())

for student in range(len(student_list)):
    student_rows = student+2
    student_name = student_list[student]
    # values
    student_overall_dedication = dedication_general_dict[student_name]
    student_specific_dedication = dedication_specific_dict[student_name]
    student_specifc_percentage = student_specific_dedication.total_seconds() \
        / dedication_specific_max.total_seconds()
    student_overall_percentage = student_overall_dedication.total_seconds() \
        / dedication_general_max.total_seconds()
    # values
    summary_sheet.cell(student_rows, 1).value = student_name
    summary_sheet.cell(student_rows, 2).value = student_specific_dedication
    summary_sheet.cell(student_rows, 3).value = student_specifc_percentage
    summary_sheet.cell(student_rows, 4).value = student_overall_dedication
    summary_sheet.cell(student_rows, 5).value = student_overall_percentage

    # formatting
    summary_sheet.cell(student_rows, 2).number_format = "[h]:mm"
    summary_sheet.cell(student_rows, 3).style = "Percent"
    summary_sheet.cell(student_rows, 3).number_format = "0.00%"
    summary_sheet.cell(student_rows, 4).number_format = "[h]:mm"
    summary_sheet.cell(student_rows, 5).style = "Percent"
    summary_sheet.cell(student_rows, 5).number_format = "0.00%"

# Get average per activity
for activity in range(len(activity_list)):
    student_total = len(student_list)+1
    activity_rows = activity + 2
    activity_name = gabarito_sheet.cell(activity_rows, 1).value
    activity_sum = activity_sum_dict[activity_name].total_seconds()
    # values
    average = activity_sum/student_total
    gabarito_sheet.cell(activity_rows, 3).value = timedelta(
        seconds=average)
    # formatting
    gabarito_sheet.cell(activity_rows, 3).number_format = "[h]:mm"

# Totals
# get values as lists
dedication_general_list = dedication_general_dict.values()

dedication_general_sec_list = []
# convert that to seconds
for value in dedication_general_list:
    value_seconds = value.total_seconds()
    dedication_general_sec_list.append(value_seconds)

gabarito_sheet.cell(gabarito_sheet.max_row, 3).value = timedelta(
    seconds=mean(dedication_general_sec_list))

# formatting
gabarito_sheet.cell(gabarito_sheet.max_row, 3).number_format = "[h]:mm"

column_fit_properly(gabarito_sheet)
column_fit_properly(summary_sheet)

# Summary Chart
# Create sheet
chart_sheet = control_book.create_sheet(title="Grafico", index=2)
# Add chart obj
activity_chart = BarChart()
activity_chart.height = 15
activity_chart.width = 30

# Add values
activity_labels = Reference(gabarito_sheet, min_col=1, min_row=2,
                            max_row=gabarito_sheet.max_row)
average_series = Reference(gabarito_sheet, min_col=3, min_row=1,
                           max_row=gabarito_sheet.max_row-1)

activity_chart.add_data(average_series, titles_from_data=True)
activity_chart.set_categories(activity_labels)
activity_chart.title = "Média por Atividade"

# place the chart
chart_sheet.add_chart(activity_chart, "A1")

# save and finish
control_book.active = control_book["Resultado"]
control_book.save(filename=os.path.basename(control_location))
summary_message = "Run Time: " + seconds_to_str(time.time() - start_time)
messagebox.showinfo("Operation Successful!", summary_message)
os.startfile(control_location)
print("\nUpdated\n")
